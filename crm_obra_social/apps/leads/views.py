import csv
import io
import re

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, UpdateView, DetailView, DeleteView

from apps.users.models import User
from apps.clientes.models import Cliente
from .forms import LeadForm, LeadFilterForm, LeadImportForm, EstadoChangeForm, CampoPersonalizadoForm
from .models import Lead, HistorialEstado, Plan, CampoPersonalizado, Documento

# Columns that are recognized as standard Lead fields (case-insensitive)
_KNOWN_COLUMNS = {
    'nombre_completo', 'nombre', 'name', 'full_name', 'apellido',
    'telefono', 'phone', 'tel', 'celular', 'movil',
    'codigo_pais', 'codigopais', 'country_code', 'cod_pais',
    'email', 'correo', 'mail',
    'dni', 'documento', 'cedula', 'rut',
    'localidad', 'ciudad', 'city',
    'provincia', 'province', 'region',
    'estado', 'status', 'estado_lead',
    'prioridad', 'priority',
    'notas', 'notes', 'observaciones', 'comentarios',
    'plan', 'plan_interes',
    'origen', 'origin', 'source', 'fuente',
    'grupo_familiar', 'grupo',
    'agente', 'agent', 'vendedor', 'asesor',
}

_ESTADO_MAP = {
    'nuevo': Lead.ESTADO_NUEVO,
    'new': Lead.ESTADO_NUEVO,
    'contactado': Lead.ESTADO_CONTACTADO,
    'contacted': Lead.ESTADO_CONTACTADO,
    'interesado': Lead.ESTADO_INTERESADO,
    'interested': Lead.ESTADO_INTERESADO,
    'doc_pendiente': Lead.ESTADO_DOC_PENDIENTE,
    'documentacion pendiente': Lead.ESTADO_DOC_PENDIENTE,
    'documentación pendiente': Lead.ESTADO_DOC_PENDIENTE,
    'en_revision': Lead.ESTADO_EN_REVISION,
    'en revision': Lead.ESTADO_EN_REVISION,
    'en revisión': Lead.ESTADO_EN_REVISION,
    'afiliado': Lead.ESTADO_AFILIADO,
    'perdido': Lead.ESTADO_PERDIDO,
    'no interesado': Lead.ESTADO_PERDIDO,
    'lost': Lead.ESTADO_PERDIDO,
}

_PRIORIDAD_MAP = {
    'alta': Lead.PRIORIDAD_ALTA,
    'high': Lead.PRIORIDAD_ALTA,
    'media': Lead.PRIORIDAD_MEDIA,
    'medium': Lead.PRIORIDAD_MEDIA,
    'normal': Lead.PRIORIDAD_MEDIA,
    'baja': Lead.PRIORIDAD_BAJA,
    'low': Lead.PRIORIDAD_BAJA,
}


def _read_file(archivo) -> tuple:
    """Read CSV or Excel. Returns (headers: list[str], rows: list[dict])."""
    name = archivo.name.lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        import openpyxl
        wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(h).strip() if h is not None else f'col_{i}' for i, h in enumerate(all_rows[0])]
        rows = []
        for raw in all_rows[1:]:
            if all(v is None or str(v).strip() == '' for v in raw):
                continue
            rows.append({headers[i]: (str(raw[i]).strip() if raw[i] is not None else '') for i in range(len(headers))})
        return headers, rows
    else:
        try:
            content = archivo.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            archivo.seek(0)
            content = archivo.read().decode('latin-1')
        reader = csv.DictReader(io.StringIO(content))
        headers = list(reader.fieldnames or [])
        rows = [dict(r) for r in reader if any(v.strip() for v in r.values())]
        return headers, rows


def _normalize_phone(raw: str, codigo_pais: str = '54') -> str:
    """Return a normalized phone like +54XXXXXXXXXX, or '' if invalid."""
    cleaned = re.sub(r'[^\d+]', '', str(raw)).strip()
    if not cleaned:
        return ''
    if cleaned.startswith('+'):
        digits = re.sub(r'\D', '', cleaned)
    else:
        digits_stripped = cleaned.lstrip('0') or cleaned
        codigo = re.sub(r'\D', '', str(codigo_pais)).lstrip('0') or '54'
        digits = codigo + digits_stripped
    if len(digits) < 7:
        return ''
    return '+' + digits


def _get(row_lower: dict, *keys, default='') -> str:
    for k in keys:
        v = row_lower.get(k.lower(), '')
        if v and str(v).strip():
            return str(v).strip()
    return default


def _resolve_agente(nombre_agente: str, agentes_cache: dict) -> object:
    """Look up an agent by full name or email, case-insensitive. Returns User or None."""
    if not nombre_agente:
        return None
    key = nombre_agente.strip().lower()
    if key in agentes_cache:
        return agentes_cache[key]
    # First call: populate cache from DB
    if not agentes_cache:
        for u in User.objects.filter(is_active=True):
            agentes_cache[u.get_full_name().lower()] = u
            agentes_cache[u.username.lower()] = u
            agentes_cache[u.email.lower()] = u
    return agentes_cache.get(key)


def _process_row(row: dict, plans_by_name: dict, actualizar: bool, default_agente, agentes_cache: dict | None = None) -> tuple:
    """Returns ('created'|'updated'|'skipped'|'error', msg, datos_extra_keys)."""
    if agentes_cache is None:
        agentes_cache = {}
    row_lower = {k.lower().strip(): v for k, v in row.items()}

    nombre = _get(row_lower, 'nombre_completo', 'nombre', 'name', 'full_name', 'apellido')
    if not nombre:
        return 'error', 'Nombre vacío', []

    raw_phone = _get(row_lower, 'telefono', 'phone', 'tel', 'celular', 'movil')
    codigo_pais = _get(row_lower, 'codigo_pais', 'codigopais', 'country_code', 'cod_pais', default='54')
    phone = _normalize_phone(raw_phone, codigo_pais)
    if not phone:
        return 'error', f'Teléfono inválido: "{raw_phone}"', []

    email = _get(row_lower, 'email', 'correo', 'mail')
    raw_dni = re.sub(r'\D', '', _get(row_lower, 'dni', 'documento', 'cedula', 'rut'))
    dni = raw_dni[:8] if raw_dni else '0000000'
    if len(dni) < 7:
        dni = dni.zfill(7)

    localidad = _get(row_lower, 'localidad', 'ciudad', 'city')
    provincia = _get(row_lower, 'provincia', 'province', 'region')
    notas = _get(row_lower, 'notas', 'notes', 'observaciones', 'comentarios')
    plan_nombre = _get(row_lower, 'plan', 'plan_interes')
    origen_raw = _get(row_lower, 'origen', 'origin', 'source', 'fuente').lower()
    origen_map = {
        'web': Lead.ORIGEN_WEB, 'campaña': Lead.ORIGEN_CAMPANA, 'campana': Lead.ORIGEN_CAMPANA,
        'referido': Lead.ORIGEN_REFERIDO, 'llamada': Lead.ORIGEN_LLAMADA, 'whatsapp': Lead.ORIGEN_WHATSAPP,
    }
    origen = origen_map.get(origen_raw, Lead.ORIGEN_CAMPANA)

    plan = plans_by_name.get(plan_nombre.lower()) if plan_nombre else None

    estado_raw = _get(row_lower, 'estado', 'status', 'estado_lead').lower()
    estado = _ESTADO_MAP.get(estado_raw, Lead.ESTADO_NUEVO)

    prioridad_raw = _get(row_lower, 'prioridad', 'priority').lower()
    prioridad = _PRIORIDAD_MAP.get(prioridad_raw, Lead.PRIORIDAD_MEDIA)

    # Agente from column takes priority over default_agente
    agente_raw = _get(row_lower, 'agente', 'agent', 'vendedor', 'asesor')
    agente = _resolve_agente(agente_raw, agentes_cache) if agente_raw else default_agente

    # Everything not recognized → datos_extra
    datos_extra = {
        k.strip(): str(v).strip()
        for k, v in row.items()
        if k.lower().strip() not in _KNOWN_COLUMNS and v and str(v).strip()
    }
    extra_keys = list(datos_extra.keys())

    # Look up duplicates: first by phone, then by DNI
    existing = Lead.objects.filter(telefono=phone).first()
    if not existing and dni and dni != '0000000':
        existing = Lead.objects.filter(dni=dni).first()

    if existing:
        if not actualizar:
            return 'skipped', '', extra_keys
        updated = []
        if not existing.nombre_completo and nombre:
            existing.nombre_completo = nombre; updated.append('nombre_completo')
        if not existing.telefono and phone:
            existing.telefono = phone; updated.append('telefono')
        if not existing.email and email:
            existing.email = email; updated.append('email')
        if not existing.localidad and localidad:
            existing.localidad = localidad; updated.append('localidad')
        if not existing.provincia and provincia:
            existing.provincia = provincia; updated.append('provincia')
        if not existing.notas and notas:
            existing.notas = notas; updated.append('notas')
        if datos_extra:
            current = existing.datos_extra or {}
            current.update(datos_extra)
            existing.datos_extra = current
            updated.append('datos_extra')
        if updated:
            existing.save(update_fields=updated + ['updated_at'])
            return 'updated', '', extra_keys
        return 'skipped', '', extra_keys

    lead = Lead.objects.create(
        nombre_completo=nombre,
        dni=dni,
        telefono=phone,
        email=email,
        localidad=localidad,
        provincia=provincia,
        notas=notas,
        origen=origen,
        plan_interes=plan,
        estado=estado,
        prioridad=prioridad,
        agente=agente,
        datos_extra=datos_extra,
    )
    HistorialEstado.objects.create(lead=lead, estado_nuevo=lead.estado, nota='Importado.')
    return 'created', '', extra_keys


class LeadQuerysetMixin:
    def get_base_queryset(self):
        qs = Lead.objects.select_related('agente', 'plan_interes').prefetch_related('conversacion_whatsapp')
        if not self.request.user.can_see_all_leads:
            qs = qs.filter(agente=self.request.user)
        return qs


class LeadListView(LoginRequiredMixin, LeadQuerysetMixin, View):
    template_name = 'leads/lead_list.html'

    def get(self, request):
        qs = self.get_base_queryset()
        form = LeadFilterForm(request.GET)

        if form.is_valid():
            data = form.cleaned_data
            if data.get('q'):
                q = data['q']
                qs = qs.filter(Q(nombre_completo__icontains=q) | Q(dni__icontains=q) | Q(telefono__icontains=q))
            if data.get('estado'):
                qs = qs.filter(estado=data['estado'])
            if data.get('plan'):
                qs = qs.filter(plan_interes=data['plan'])
            if data.get('provincia'):
                qs = qs.filter(provincia__icontains=data['provincia'])
            if data.get('origen'):
                qs = qs.filter(origen=data['origen'])
            if data.get('prioridad'):
                qs = qs.filter(prioridad=data['prioridad'])
            if data.get('fecha_desde'):
                qs = qs.filter(created_at__date__gte=data['fecha_desde'])
            if data.get('fecha_hasta'):
                qs = qs.filter(created_at__date__lte=data['fecha_hasta'])
            if data.get('agente') and request.user.can_see_all_leads:
                qs = qs.filter(agente_id=data['agente'])

        paginator = Paginator(qs, 25)
        page = paginator.get_page(request.GET.get('page'))

        agents = User.objects.filter(is_active=True).order_by('first_name') if request.user.can_see_all_leads else None

        return render(request, self.template_name, {
            'leads': page,
            'filter_form': form,
            'agents': agents,
            'total_count': qs.count(),
        })


class LeadKanbanView(LoginRequiredMixin, LeadQuerysetMixin, View):
    template_name = 'leads/lead_kanban.html'

    _ESTADO_COLORS = {
        'nuevo': 'secondary', 'contactado': 'info', 'interesado': 'primary',
        'doc_pendiente': 'warning', 'en_revision': 'warning',
        'afiliado': 'success', 'perdido': 'danger',
    }

    def get(self, request):
        qs = self.get_base_queryset()
        columns = {}
        for estado_key, estado_label in Lead.ESTADO_CHOICES:
            columns[estado_key] = {
                'label': estado_label,
                'color': self._ESTADO_COLORS.get(estado_key, 'secondary'),
                'leads': qs.filter(estado=estado_key).order_by('-prioridad', '-updated_at')[:50],
            }
        return render(request, self.template_name, {'columns': columns, 'estados': Lead.ESTADO_CHOICES})


class LeadCreateView(LoginRequiredMixin, CreateView):
    model = Lead
    form_class = LeadForm
    template_name = 'leads/lead_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['from_contacts'] = self.request.GET.get('from', '')
        return ctx

    def form_valid(self, form):
        lead = form.save(commit=False)
        if not lead.agente and not self.request.user.can_see_all_leads:
            lead.agente = self.request.user
        lead.save()
        HistorialEstado.objects.create(
            lead=lead,
            estado_nuevo=lead.estado,
            cambiado_por=self.request.user,
            nota='Lead creado.',
        )
        messages.success(self.request, 'Contacto creado correctamente.')
        if self.request.POST.get('from') == 'contacts':
            return redirect('leads:contact_list')
        return redirect('leads:detail', pk=lead.pk)


class LeadDetailView(LoginRequiredMixin, LeadQuerysetMixin, DetailView):
    template_name = 'leads/lead_detail.html'
    context_object_name = 'lead'

    def get_object(self):
        return get_object_or_404(self.get_base_queryset(), pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        lead = self.object
        ctx['historial'] = lead.historial_estados.select_related('cambiado_por').all()
        ctx['estado_form'] = EstadoChangeForm(initial={'estado': lead.estado})
        ctx['tareas'] = lead.tareas.select_related('agente').order_by('-fecha_programada')[:10]
        ctx['cotizaciones'] = lead.cotizaciones.order_by('-created_at')[:5]
        ctx['mensajes'] = lead.mensajes_whatsapp.order_by('-timestamp')[:20]
        ctx['campos'] = CampoPersonalizado.objects.filter(
            activo=True,
            alcance__in=[CampoPersonalizado.ALCANCE_LEADS, CampoPersonalizado.ALCANCE_AMBOS]
        )
        ctx['documentos'] = lead.documentos.select_related('subido_por').all()
        ctx['documento_tipos'] = Documento.TIPO_CHOICES
        return ctx


class LeadUpdateView(LoginRequiredMixin, LeadQuerysetMixin, UpdateView):
    model = Lead
    form_class = LeadForm
    template_name = 'leads/lead_form.html'

    def get_object(self):
        return get_object_or_404(self.get_base_queryset(), pk=self.kwargs['pk'])

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['from_contacts'] = self.request.GET.get('from', '')
        return ctx

    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'Lead actualizado correctamente.')
        return redirect('leads:detail', pk=self.object.pk)


class LeadDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Lead
    template_name = 'leads/lead_confirm_delete.html'
    success_url = reverse_lazy('leads:list')

    def test_func(self):
        return self.request.user.can_see_all_leads

    def form_valid(self, form):
        messages.success(self.request, 'Lead eliminado.')
        return super().form_valid(form)


class LeadUpdateCamposView(LoginRequiredMixin, LeadQuerysetMixin, View):
    """Save custom field values for a lead."""

    def post(self, request, pk):
        lead = get_object_or_404(self.get_base_queryset(), pk=pk)
        campos = CampoPersonalizado.objects.filter(
            activo=True,
            alcance__in=[CampoPersonalizado.ALCANCE_LEADS, CampoPersonalizado.ALCANCE_AMBOS]
        )
        extra = dict(lead.datos_extra or {})
        for campo in campos:
            if campo.tipo == CampoPersonalizado.TIPO_BOOLEANO:
                extra[campo.slug] = bool(request.POST.get(f'campo_{campo.slug}'))
            else:
                val = request.POST.get(f'campo_{campo.slug}', '').strip()
                if val:
                    extra[campo.slug] = val
                else:
                    extra.pop(campo.slug, None)
        lead.datos_extra = extra
        lead.save(update_fields=['datos_extra'])
        messages.success(request, 'Campos guardados.')
        return redirect('leads:detail', pk=pk)


class LeadConvertirView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Convert a lead to a cliente (separate model). Deletes the lead."""

    def test_func(self):
        return self.request.user.can_see_all_leads

    def post(self, request, pk):
        lead = get_object_or_404(Lead, pk=pk)
        from apps.clientes.models import Cliente
        cliente = Cliente.objects.create(
            nombre_completo=lead.nombre_completo,
            dni=lead.dni,
            fecha_nacimiento=lead.fecha_nacimiento,
            telefono=lead.telefono,
            email=lead.email,
            localidad=lead.localidad,
            provincia=lead.provincia,
            plan=lead.plan_interes,
            grupo_familiar=lead.grupo_familiar,
            agente=lead.agente,
            notas=lead.notas,
            datos_extra=lead.datos_extra or {},
        )
        lead.delete()
        messages.success(request, f'{cliente.nombre_completo} convertido a Cliente correctamente.')
        return redirect('clientes:detail', pk=cliente.pk)


# ── Documentos ──────────────────────────────────────────

class DocumentoUploadView(LoginRequiredMixin, View):
    def post(self, request, lead_pk):
        lead = get_object_or_404(Lead, pk=lead_pk)
        if not request.user.can_see_all_leads and lead.agente != request.user:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()
        archivo = request.FILES.get('archivo')
        nombre = request.POST.get('nombre', '').strip()
        tipo = request.POST.get('tipo', Documento.TIPO_OTRO)
        if not archivo:
            messages.error(request, 'Seleccioná un archivo.')
            return redirect('leads:detail', pk=lead_pk)
        if archivo.size > 20 * 1024 * 1024:
            messages.error(request, 'El archivo no puede superar 20 MB.')
            return redirect('leads:detail', pk=lead_pk)
        Documento.objects.create(
            lead=lead,
            nombre=nombre or archivo.name,
            tipo=tipo,
            archivo=archivo,
            subido_por=request.user,
        )
        messages.success(request, 'Documento subido correctamente.')
        return redirect('leads:detail', pk=lead_pk)


class DocumentoDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        doc = get_object_or_404(Documento, pk=pk)
        lead_pk = doc.lead_id
        if not request.user.can_see_all_leads and doc.lead.agente != request.user:
            from django.http import HttpResponseForbidden
            return HttpResponseForbidden()
        doc.archivo.delete(save=False)
        doc.delete()
        messages.success(request, 'Documento eliminado.')
        return redirect('leads:detail', pk=lead_pk)


# ── Campos personalizados CRUD ────────────────────────────

class SupervisorRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.can_see_all_leads


class CampoListView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'leads/campo_list.html'

    def get(self, request):
        campos = CampoPersonalizado.objects.all()
        return render(request, self.template_name, {'campos': campos})


class CampoCreateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'leads/campo_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': CampoPersonalizadoForm()})

    def post(self, request):
        form = CampoPersonalizadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo creado.')
            return redirect('leads:campos')
        return render(request, self.template_name, {'form': form})


class CampoUpdateView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    template_name = 'leads/campo_form.html'

    def get(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        return render(request, self.template_name, {'form': CampoPersonalizadoForm(instance=campo), 'campo': campo})

    def post(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        form = CampoPersonalizadoForm(request.POST, instance=campo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Campo actualizado.')
            return redirect('leads:campos')
        return render(request, self.template_name, {'form': form, 'campo': campo})


class CampoDeleteView(LoginRequiredMixin, SupervisorRequiredMixin, DeleteView):
    model = CampoPersonalizado
    template_name = 'leads/campo_confirm_delete.html'
    success_url = reverse_lazy('leads:campos')


class CampoToggleView(LoginRequiredMixin, SupervisorRequiredMixin, View):
    def post(self, request, pk):
        campo = get_object_or_404(CampoPersonalizado, pk=pk)
        campo.activo = not campo.activo
        campo.save(update_fields=['activo'])
        return redirect('leads:campos')


class LeadEstadoChangeView(LoginRequiredMixin, LeadQuerysetMixin, View):
    def post(self, request, pk):
        lead = get_object_or_404(self.get_base_queryset(), pk=pk)
        form = EstadoChangeForm(request.POST)
        if form.is_valid():
            estado_anterior = lead.estado
            lead.estado = form.cleaned_data['estado']
            motivo = form.cleaned_data.get('motivo_perdida', '').strip()
            update_fields = ['estado', 'updated_at']
            if motivo:
                lead.motivo_perdida = motivo
                update_fields.append('motivo_perdida')
            lead.save(update_fields=update_fields)
            HistorialEstado.objects.create(
                lead=lead,
                estado_anterior=estado_anterior,
                estado_nuevo=lead.estado,
                cambiado_por=request.user,
                nota=form.cleaned_data.get('nota', ''),
            )
            messages.success(request, f'Estado cambiado a "{lead.get_estado_display()}".')
        else:
            for field, errs in form.errors.items():
                for err in errs:
                    messages.error(request, err)
        return redirect('leads:detail', pk=pk)


class LeadKanbanMoveView(LoginRequiredMixin, LeadQuerysetMixin, View):
    """AJAX endpoint for Kanban drag & drop."""

    def post(self, request, pk):
        lead = get_object_or_404(self.get_base_queryset(), pk=pk)
        nuevo_estado = request.POST.get('estado')
        valid_states = dict(Lead.ESTADO_CHOICES).keys()
        if nuevo_estado not in valid_states:
            return JsonResponse({'error': 'Estado inválido.'}, status=400)
        estado_anterior = lead.estado
        lead.estado = nuevo_estado
        lead.save(update_fields=['estado', 'updated_at'])
        HistorialEstado.objects.create(
            lead=lead,
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            cambiado_por=request.user,
        )
        return JsonResponse({'ok': True, 'estado': nuevo_estado, 'estado_display': lead.get_estado_display()})


class ContactListView(LoginRequiredMixin, View):
    """Combined contact list: Leads + Clientes merged and sorted by name."""
    template_name = 'leads/contact_list.html'

    def get(self, request):
        q = request.GET.get('q', '').strip()
        plan_id = request.GET.get('plan')
        tipo_filter = request.GET.get('tipo', '')

        leads_qs = Lead.objects.select_related('plan_interes', 'agente').prefetch_related('conversacion_whatsapp')
        clientes_qs = Cliente.objects.select_related('plan', 'agente')

        if q:
            leads_qs = leads_qs.filter(
                Q(nombre_completo__icontains=q) | Q(telefono__icontains=q) | Q(email__icontains=q)
            )
            clientes_qs = clientes_qs.filter(
                Q(nombre_completo__icontains=q) | Q(telefono__icontains=q) | Q(email__icontains=q)
            )

        if plan_id:
            leads_qs = leads_qs.filter(plan_interes_id=plan_id)
            clientes_qs = clientes_qs.filter(plan_id=plan_id)

        # Tag each object so the template can distinguish them
        leads_list = list(leads_qs)
        for obj in leads_list:
            obj.tipo_contacto = 'lead'

        clientes_list = list(clientes_qs)
        for obj in clientes_list:
            obj.tipo_contacto = 'cliente'

        if tipo_filter == 'lead':
            combined = leads_list
        elif tipo_filter == 'cliente':
            combined = clientes_list
        else:
            combined = leads_list + clientes_list

        combined.sort(key=lambda c: c.nombre_completo.lower())

        paginator = Paginator(combined, 30)
        page = paginator.get_page(request.GET.get('page'))

        return render(request, self.template_name, {
            'contactos': page,
            'q': q,
            'plan_id': plan_id,
            'tipo_filter': tipo_filter,
            'planes': Plan.objects.filter(activo=True),
            'total': len(combined),
        })


class ContactSearchAPIView(LoginRequiredMixin, View):
    """AJAX: search contacts for quick lookup."""

    def get(self, request):
        from django.db.models import Q
        q = request.GET.get('q', '').strip()
        qs = Lead.objects.filter(telefono__startswith='+').order_by('nombre_completo')
        if q:
            qs = qs.filter(
                Q(nombre_completo__icontains=q) |
                Q(telefono__icontains=q) |
                Q(email__icontains=q)
            )
        results = [
            {
                'id': l.pk,
                'nombre': l.nombre_completo,
                'telefono': l.telefono,
                'email': l.email or '',
            }
            for l in qs[:20]
        ]
        return JsonResponse({'results': results})


class LeadCSVImportView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = 'leads/lead_import.html'

    def test_func(self):
        return self.request.user.can_see_all_leads

    def get(self, request):
        return render(request, self.template_name, {'form': LeadImportForm()})

    def post(self, request):
        form = LeadImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        archivo = request.FILES['archivo']
        actualizar = form.cleaned_data.get('actualizar_existentes', True)
        asignar = form.cleaned_data.get('asignar_agente', False)
        default_agente = request.user if asignar else None

        try:
            headers, rows = _read_file(archivo)
        except Exception as e:
            messages.error(request, f'Error leyendo el archivo: {e}')
            return render(request, self.template_name, {'form': form})

        if not rows:
            messages.warning(request, 'El archivo está vacío o no tiene filas de datos.')
            return render(request, self.template_name, {'form': form})

        plans_by_name = {p.nombre.lower(): p for p in Plan.objects.filter(activo=True)}
        plans_by_name.update({p.nombre: p for p in Plan.objects.filter(activo=True)})

        # Pre-load all agents once for the whole import
        agentes_cache = {}
        for u in User.objects.filter(is_active=True):
            agentes_cache[u.get_full_name().lower()] = u
            agentes_cache[u.username.lower()] = u
            agentes_cache[u.email.lower()] = u

        stats = {'created': 0, 'updated': 0, 'skipped': 0, 'error': 0}
        errors = []
        all_extra_keys = set()

        for i, row in enumerate(rows, start=2):
            action, msg, extra_keys = _process_row(row, plans_by_name, actualizar, default_agente, agentes_cache)
            stats[action] += 1
            all_extra_keys.update(extra_keys)
            if action == 'error':
                errors.append({'fila': i, 'msg': msg, 'datos': str(row)[:120]})

        ctx = {
            'form': LeadImportForm(),
            'resultado': True,
            'stats': stats,
            'errors': errors[:50],
            'total_filas': len(rows),
            'extra_keys': sorted(all_extra_keys),
        }
        return render(request, self.template_name, ctx)


class LeadCSVExportView(LoginRequiredMixin, LeadQuerysetMixin, View):
    def get(self, request):
        qs = self.get_base_queryset()
        # Collect all datos_extra keys across all leads
        extra_keys = set()
        for lead in qs:
            extra_keys.update((lead.datos_extra or {}).keys())
        extra_keys = sorted(extra_keys)

        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="leads.csv"'
        response.write('﻿')
        writer = csv.writer(response)
        base_headers = ['ID', 'nombre_completo', 'dni', 'telefono', 'email', 'localidad', 'provincia',
                        'plan', 'estado', 'prioridad', 'origen', 'notas', 'agente', 'creado']
        writer.writerow(base_headers + extra_keys)
        for lead in qs:
            base = [
                lead.pk, lead.nombre_completo, lead.dni, lead.telefono, lead.email,
                lead.localidad, lead.provincia,
                lead.plan_interes.nombre if lead.plan_interes else '',
                lead.get_estado_display(), lead.get_prioridad_display(), lead.get_origen_display(),
                lead.notas,
                lead.agente.get_full_name() if lead.agente else '',
                lead.created_at.strftime('%d/%m/%Y %H:%M'),
            ]
            extras = [(lead.datos_extra or {}).get(k, '') for k in extra_keys]
            writer.writerow(base + extras)
        return response


class LeadImportTemplateView(LoginRequiredMixin, View):
    """Download an Excel import template with column explanations."""

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Contactos'

        headers = [
            'nombre_completo', 'codigo_pais', 'telefono', 'email', 'dni',
            'localidad', 'provincia', 'plan', 'notas', 'origen',
            'empresa', 'cargo',
        ]
        descriptions = [
            'Nombre y apellido *', 'Cód. de país (ej: 54)', 'Teléfono sin código de país *',
            'Email', 'DNI (7-8 dígitos)', 'Localidad / Ciudad', 'Provincia',
            'Nombre exacto del plan', 'Notas internas',
            'web / campaña / referido / llamada / whatsapp',
            'Columna extra de ejemplo', 'Otra columna extra',
        ]

        header_fill = PatternFill('solid', fgColor='1E3A5F')
        desc_fill = PatternFill('solid', fgColor='E8F0FE')
        extra_fill = PatternFill('solid', fgColor='FFF3CD')
        bold_white = Font(bold=True, color='FFFFFF')
        bold_dark = Font(bold=True, color='333333')

        for col, (header, desc) in enumerate(zip(headers, descriptions), start=1):
            is_extra = col > 10
            # Header row
            hcell = ws.cell(row=1, column=col, value=header)
            hcell.font = bold_white
            hcell.fill = extra_fill if is_extra else header_fill
            hcell.alignment = Alignment(horizontal='center')
            # Description row
            dcell = ws.cell(row=2, column=col, value=desc)
            dcell.fill = PatternFill('solid', fgColor='FFF3CD') if is_extra else desc_fill
            dcell.font = Font(italic=True, color='666666' if not is_extra else '856404')
            ws.column_dimensions[get_column_letter(col)].width = max(len(header), len(desc)) + 4

        # Example rows
        examples = [
            ['Juan Pérez', '54', '1123456789', 'juan@gmail.com', '35123456', 'Buenos Aires', 'Buenos Aires', 'Plan Individual', '', 'web', 'Empresa SA', 'Gerente'],
            ['María García', '54', '2996543210', 'maria@gmail.com', '', 'Neuquén', 'Neuquén', '', 'Interesada en familiar', 'whatsapp', '', ''],
        ]
        for r, row in enumerate(examples, start=3):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)

        # Instruction sheet
        ws2 = wb.create_sheet('Instrucciones')
        instructions = [
            ('INSTRUCCIONES DE IMPORTACIÓN', True),
            ('', False),
            ('COLUMNAS REQUERIDAS (* obligatorias):', True),
            ('  nombre_completo — Nombre y apellido del contacto', False),
            ('  telefono — Solo el número, sin código de país ni espacios ni guiones', False),
            ('', False),
            ('CÓDIGO DE PAÍS:', True),
            ('  codigo_pais — Ej: 54 (Argentina), 598 (Uruguay), 56 (Chile)', False),
            ('  Si no se incluye esta columna, se asume 54 (Argentina)', False),
            ('', False),
            ('COLUMNAS ESTÁNDAR (opcionales):', True),
            ('  email, dni, localidad, provincia, plan, notas, origen', False),
            ('  origen válidos: web, campaña, referido, llamada, whatsapp', False),
            ('', False),
            ('COLUMNAS EXTRAS (datos personalizados):', True),
            ('  Podés agregar cualquier otra columna con el nombre que quieras.', False),
            ('  Ej: empresa, cargo, fecha_cumpleaños, numero_socio', False),
            ('  Estos datos se guardan en el lead y pueden usarse como variables', False),
            ('  en las plantillas de WhatsApp al crear campañas masivas.', False),
            ('', False),
            ('COMPORTAMIENTO DEL IMPORT:', True),
            ('  - Si un número de teléfono ya existe → se actualiza el contacto', False),
            ('  - Si el número no existe → se crea un lead nuevo', False),
            ('  - Teléfonos duplicados dentro del archivo → se procesa el primero', False),
            ('', False),
            ('FORMATOS DE TELÉFONO ACEPTADOS:', True),
            ('  1123456789  →  +541123456789', False),
            ('  011 2345-6789  →  +541123456789', False),
            ('  +54 11 2345-6789  →  +541123456789', False),
        ]
        for row_idx, (text, is_bold) in enumerate(instructions, start=1):
            cell = ws2.cell(row=row_idx, column=1, value=text)
            if is_bold:
                cell.font = Font(bold=True)
        ws2.column_dimensions['A'].width = 65

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template_importar_contactos.xlsx"'
        wb.save(response)
        return response


class LeadBulkAssignView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.can_see_all_leads

    def post(self, request):
        lead_ids = request.POST.getlist('lead_ids')
        agente_id = request.POST.get('agente_id', '').strip()

        if not lead_ids:
            messages.warning(request, 'No seleccionaste ningún lead.')
            return redirect(request.POST.get('next', 'leads:list'))

        try:
            agente = User.objects.get(pk=agente_id, is_active=True)
        except (User.DoesNotExist, ValueError):
            messages.error(request, 'Agente inválido.')
            return redirect(request.POST.get('next', 'leads:list'))

        updated = Lead.objects.filter(pk__in=lead_ids).update(agente=agente)
        messages.success(request, f'{updated} lead{"s" if updated != 1 else ""} asignado{"s" if updated != 1 else ""} a {agente.get_full_name()}.')
        return redirect(request.POST.get('next', 'leads:list'))
